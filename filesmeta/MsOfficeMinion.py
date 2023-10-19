# pip install python-docx
import docx
# pip install python-pptx
import pptx
from pptx import Presentation
from openpyxl import load_workbook

class MsOfficeMinion():

    ex = ['docx', 'pptx', 'xlsx']

    def get_extension(path_to_a_file):
        last_part = path_to_a_file.split("/")[-1]
        if "." in last_part:
            ex = path_to_a_file.split(".")[-1]
        else:
            ex = "У файла нет расширения"
        return ex

    def get_meta_inf(self, path):
        if get_extension(path) == 'docx':
            try:
                fname = path
                doc = docx.Document(fname)

                prop = doc.core_properties

                data_docx = {}
                for d in dir(prop):
                    if not d.startswith('_'):
                        data_docx[d] = getattr(prop, d)



                return data_docx

            except:
                print(f"{path} is broken and cannot be read!")
                return None

        elif get_extension(path) == 'pptx':
            try:
                fname = path
                pres = pptx.Presentation(fname)

                prop = pres.core_properties

                data_pptx = {}
                for d in dir(prop):
                    if not d.startswith('_'):
                        data_pptx[d] = getattr(prop, d)

                num_of_slides = {"Slides": len(pres.slides)}

                data_pptx.update(num_of_slides)

                return data_pptx

            except:
                print(f"{path} is broken and cannot be read!")
                return None


        elif get_extension(path) == 'xlsx':
            try:

                fname = path
                table = load_workbook(fname)

                prop = table.properties

                data_xlsx = {}
                for d in dir(prop):
                    if not d.startswith('_'):
                        data_xlsx[d] = getattr(prop, d)

                num_of_sheets = {"Sheets": len(table.sheetnames)}

                data_xlsx.update(num_of_sheets)

                return data_xlsx

            except:
                print(f"{path} is broken and cannot be read!")
                return None
